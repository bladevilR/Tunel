from __future__ import annotations

import json
import re
import zipfile
from pathlib import Path
from statistics import mean
from xml.etree import ElementTree as ET

from docx import Document
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKSPACE_ROOT = ROOT.parents[1]
FEEDBACK_DIR = ROOT.parent / "feedback_20260507"
DATA_DIR = ROOT / "data"
DOCS_DIR = ROOT / "docs"


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def as_number(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def month_key(value: object) -> str | None:
    text = clean(value)
    if not re.fullmatch(r"\d{6}", text):
        return None
    return f"{text[:4]}-{text[4:]}"


def write_json(name: str, payload: dict) -> None:
    target = DATA_DIR / name
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def find_feedback_file(token: str, suffix: str) -> Path:
    matches = [path for path in FEEDBACK_DIR.rglob(f"*{suffix}") if token in path.name and not path.name.startswith("~$")]
    if not matches:
        raise FileNotFoundError(f"Missing feedback file containing {token!r} with suffix {suffix!r}")
    return sorted(matches, key=lambda path: path.stat().st_size)[0]


def station_aliases(name: str) -> list[str]:
    text = clean(name)
    aliases = [text]
    aliases.extend(match.strip() for match in re.findall(r"[（(]([^）)]+)[）)]", text) if match.strip())
    no_paren = re.sub(r"[（(].*?[）)]", "", text).strip()
    if no_paren:
        aliases.append(no_paren)
    if no_paren.endswith("站"):
        aliases.append(no_paren[:-1])
    return list(dict.fromkeys(item for item in aliases if item))


def station_key(name: str) -> str:
    aliases = station_aliases(name)
    if len(aliases) > 1:
        return aliases[1]
    return aliases[0] if aliases else clean(name)


def parse_ridership() -> dict:
    source = find_feedback_file("日均进站", ".xlsx")
    wb = load_workbook(source, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    month_row = rows[1]
    month_pairs: list[tuple[int, int, str]] = []
    for index, value in enumerate(header):
        if clean(value) == "累计进站":
            month = month_key(month_row[index])
            if month and index + 1 < len(header) and clean(header[index + 1]) == "日均进站":
                month_pairs.append((index, index + 1, month))

    records = []
    for row in rows[2:]:
        station_name = clean(row[2] if len(row) > 2 else "")
        if not station_name:
            continue
        monthly = []
        for total_col, daily_col, month in month_pairs:
            total = as_number(row[total_col] if total_col < len(row) else None)
            daily = as_number(row[daily_col] if daily_col < len(row) else None)
            if total is not None or daily is not None:
                monthly.append({
                    "month": month,
                    "inboundTotal": round(total, 2) if total is not None else None,
                    "dailyInbound": round(daily, 2) if daily is not None else None
                })
        daily_values = [item["dailyInbound"] for item in monthly if item.get("dailyInbound") is not None]
        latest = next((item for item in reversed(monthly) if item.get("dailyInbound") is not None), None)
        records.append({
            "stationName": station_name,
            "lines": clean(row[1] if len(row) > 1 else ""),
            "latestMonth": latest["month"] if latest else None,
            "latestDailyInbound": latest["dailyInbound"] if latest else None,
            "averageDailyInbound": round(mean(daily_values), 2) if daily_values else None,
            "monthCount": len(monthly),
            "monthly": monthly
        })
    wb.close()
    return {
        "version": "feedback-20260507",
        "source": str(source.relative_to(WORKSPACE_ROOT)),
        "metric": "日均进站",
        "count": len(records),
        "records": records
    }


def column_index(headers: list[object], token: str) -> int | None:
    for index, header in enumerate(headers):
        if token in clean(header):
            return index
    return None


def status_columns(header_following_row: list[object]) -> dict[str, int]:
    status = {}
    for index, value in enumerate(header_following_row):
        text = clean(value)
        if text in {"已开通", "建设中", "已建设未开通", "未建设未开通"}:
            status[text] = index
    return status


def check_mark(value: object) -> bool:
    return clean(value) in {"√", "✓", "1", "是"}


def merge_set(target: dict, key: str, value: str) -> None:
    if value:
        target.setdefault(key, set()).add(value)


def parse_station_operations() -> dict:
    source = find_feedback_file("出入口", ".xlsx")
    wb = load_workbook(source, read_only=True, data_only=True)
    station_map: dict[str, dict] = {}
    detail_count = 0

    for ws in wb.worksheets:
        raw_rows = list(ws.iter_rows(values_only=True))
        header_row = None
        for index, row in enumerate(raw_rows[:8]):
            values = [clean(item) for item in row]
            if "站点名称" in values and "部位" in values:
                header_row = index
                break
        if header_row is None:
            continue

        headers = list(raw_rows[header_row])
        status_map = status_columns(list(raw_rows[header_row + 1]) if header_row + 1 < len(raw_rows) else [])
        col_line = column_index(headers, "线路")
        col_station = column_index(headers, "站点名称")
        col_district = column_index(headers, "所属区划")
        col_type = column_index(headers, "类型")
        col_part = column_index(headers, "部位")
        col_parcel = column_index(headers, "地块")
        col_form = column_index(headers, "联通形式")
        col_progress = column_index(headers, "开发进度")
        col_problem = column_index(headers, "存在问题")
        col_priority = column_index(headers, "优先分级")

        current: dict[str, str] = {}
        for row in raw_rows[header_row + 2:]:
            if not any(clean(value) for value in row):
                continue

            for key, col in {
                "line": col_line,
                "station": col_station,
                "district": col_district,
                "type": col_type,
            }.items():
                if col is not None and col < len(row) and clean(row[col]):
                    current[key] = clean(row[col])

            raw_station = current.get("station", "")
            if not raw_station:
                continue

            key = station_key(raw_station)
            record = station_map.setdefault(key, {
                "name": key,
                "displayNames": set(),
                "aliases": set(),
                "lines": set(),
                "districts": set(),
                "types": set(),
                "exitParts": set(),
                "connectionForms": set(),
                "statusCounts": {"已开通": 0, "建设中": 0, "已建设未开通": 0, "未建设未开通": 0},
                "issueCount": 0,
                "priorityLabels": set(),
                "details": []
            })
            record["displayNames"].add(raw_station)
            for alias in station_aliases(raw_station):
                record["aliases"].add(alias)
            merge_set(record, "lines", current.get("line", ""))
            merge_set(record, "districts", current.get("district", ""))
            merge_set(record, "types", current.get("type", ""))

            part = clean(row[col_part]) if col_part is not None and col_part < len(row) else ""
            parcel = clean(row[col_parcel]) if col_parcel is not None and col_parcel < len(row) else ""
            form = clean(row[col_form]) if col_form is not None and col_form < len(row) else ""
            progress = clean(row[col_progress]) if col_progress is not None and col_progress < len(row) else ""
            problem = clean(row[col_problem]) if col_problem is not None and col_problem < len(row) else ""
            priority = clean(row[col_priority]) if col_priority is not None and col_priority < len(row) else ""
            status = next((label for label, col in status_map.items() if col < len(row) and check_mark(row[col])), "")

            if part and ("号口" in part or "出入口" in current.get("type", "")):
                record["exitParts"].add(part)
            if form and form not in {"/", "-", "无"}:
                record["connectionForms"].add(form)
            if status:
                record["statusCounts"][status] += 1
            if problem:
                record["issueCount"] += 1
            if priority:
                record["priorityLabels"].add(priority)

            if (part or parcel or form or status or problem) and len(record["details"]) < 30:
                record["details"].append({
                    "line": current.get("line", ""),
                    "displayStationName": raw_station,
                    "district": current.get("district", ""),
                    "type": current.get("type", ""),
                    "part": part,
                    "parcel": parcel,
                    "connectionForm": form,
                    "developmentProgress": progress,
                    "status": status,
                    "problem": problem,
                    "priority": priority
                })
                detail_count += 1

    wb.close()

    records = []
    for record in station_map.values():
        exit_parts = sorted(record.pop("exitParts"))
        forms = sorted(record["connectionForms"])
        records.append({
            **record,
            "displayNames": sorted(record["displayNames"]),
            "aliases": sorted(record["aliases"]),
            "lines": sorted(record["lines"]),
            "districts": sorted(record["districts"]),
            "types": sorted(record["types"]),
            "exitCount": len(exit_parts),
            "exitParts": exit_parts,
            "interfaceCount": len(forms),
            "connectionForms": forms,
            "priorityLabels": sorted(record["priorityLabels"]),
        })
    records.sort(key=lambda item: item["name"])
    return {
        "version": "feedback-20260507",
        "source": str(source.relative_to(WORKSPACE_ROOT)),
        "count": len(records),
        "detailRowsSampled": detail_count,
        "records": records
    }


def parse_input_schema() -> dict:
    source = find_feedback_file("标准输入要素", ".docx")
    doc = Document(source)
    sections: dict[str, list[str]] = {}
    current = "未分组"
    for paragraph in doc.paragraphs:
        text = paragraph.text.strip()
        if not text:
            continue
        if "：" not in text and len(text) <= 14:
            current = text
            sections.setdefault(current, [])
        else:
            sections.setdefault(current, []).append(text)

    fields = [
        {"path": "station.district", "label": "所属区划", "required": True, "example": "工业园区"},
        {"path": "station.name", "label": "轨道站点", "required": True, "example": "金家堰站"},
        {"path": "name", "label": "研究地块", "required": True, "example": "金家堰邻里中心"},
        {"path": "parcel.location", "label": "地块位置", "required": True, "example": "金堰路东、东宏路北"},
        {"path": "station.nearbyExit", "label": "车站相邻出入口", "required": True, "example": "3号口"},
        {"path": "parcel.landUseText", "label": "用地性质", "required": True, "example": "商业用地"},
        {"path": "parcel.siteArea", "label": "用地面积", "required": False, "unit": "平方米", "example": "25000"},
        {"path": "parcel.functionalFormat", "label": "功能业态", "required": True, "example": "综合性社区邻里中心"},
        {"path": "parcel.undergroundDescription", "label": "地下空间", "required": True, "example": "有地下室"},
        {"path": "parcel.planningIndicators", "label": "规划指标", "required": False, "example": "容积率≤2.5，建筑密度≤50%，建筑限高60m"},
        {"path": "attachments", "label": "附件资料", "required": False, "example": "控规图则、规划条件、概念方案"}
    ]
    return {
        "version": "feedback-20260507",
        "source": str(source.relative_to(WORKSPACE_ROOT)),
        "sections": sections,
        "fields": fields
    }


def parse_ppt_summary() -> dict:
    source = find_feedback_file("0318", ".pptx")
    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    slides = []
    with zipfile.ZipFile(source) as zf:
        names = sorted(
            [name for name in zf.namelist() if re.match(r"ppt/slides/slide\d+\.xml$", name)],
            key=lambda name: int(re.search(r"(\d+)", name).group(1))
        )
        for index, name in enumerate(names, 1):
            tree = ET.fromstring(zf.read(name))
            text = " ".join(node.text.strip() for node in tree.findall(".//a:t", ns) if node.text and node.text.strip())
            slides.append({"slide": index, "text": text})
    return {
        "version": "feedback-20260507",
        "source": str(source.relative_to(WORKSPACE_ROOT)),
        "priority": "PPT is the authoritative work-plan source; old example reports are low-priority references.",
        "mvpGoal": "标准化联通必要性评估报告和联通设计方案建议报告（文字版本）",
        "grading": "必连 / 尽连 / 可连三档；20260507反馈暂按80分以上、60-80分、60分以下。",
        "outputTemplateSections": [
            "项目概况",
            "联通必要性评估结论",
            "联通方式比选与推荐方案",
            "方案设计核心技术要求",
            "合规性要求说明",
            "设计优化与实施建议"
        ],
        "slides": slides
    }


def build_feedback_note(payloads: dict[str, dict]) -> None:
    target = DOCS_DIR / "feedback_20260507_analysis.md"
    ridership = payloads["ridership"]
    operations = payloads["operations"]
    schema = payloads["input_schema"]
    target.write_text(
        "\n".join([
            "# 20260507反馈提资解析",
            "",
            "## 已明确口径",
            "",
            "- 评分口径以 `评价因子赋值明细表.xlsx` 为准。",
            "- 用地功能允许 5 分；一般站按 2 分。",
            "- 客流一期先采用 `每站每月日均进站.xlsx` 的日均进站口径，后续可替换预测客流。",
            "- 地块只分三级：80分以上必连、60-80分尽连、60分以下可连，取消不连地块。",
            "- 一期成果边界为文字报告与意向效果简图；CAD/BIM自动审查不纳入一期。",
            "- 标准输入界面参照 `标准输入要素.docx`。",
            "",
            "## 已转入系统的数据",
            "",
            f"- 客流记录：{ridership['count']} 个站点，指标为 `{ridership['metric']}`。",
            f"- 出入口/联通接口摘要：{operations['count']} 个站点，保留出入口数、联通形式、建设状态、问题与优先级摘要。",
            f"- 标准输入字段：{len(schema['fields'])} 个字段，覆盖区位、站点、地块、地下空间、规划指标与附件。",
            "",
            "## 系统处理策略",
            "",
            "- 前端站点选择后可自动带出 TOD、线路、区位能级和客流；人工输入仍可覆盖。",
            "- 后端评分按原始加权分计算，同时换算为百分制后按 80/60 阈值判定等级。",
            "- 出入口与接口表先作为站点侧上下文进入报告，不把该指标硬并入评分，符合反馈中“影响较小、可后续完善”的口径。",
            "",
        ]),
        encoding="utf-8"
    )


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    payloads = {
        "ridership": parse_ridership(),
        "operations": parse_station_operations(),
        "input_schema": parse_input_schema(),
        "ppt_rules": parse_ppt_summary(),
    }
    write_json("ridership.json", payloads["ridership"])
    write_json("station_operations.json", payloads["operations"])
    write_json("input_schema.json", payloads["input_schema"])
    write_json("ppt_rules_summary.json", payloads["ppt_rules"])
    build_feedback_note(payloads)
    print(json.dumps({
        "ridership": payloads["ridership"]["count"],
        "operations": payloads["operations"]["count"],
        "inputFields": len(payloads["input_schema"]["fields"]),
        "docs": str((DOCS_DIR / "feedback_20260507_analysis.md").relative_to(ROOT))
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
